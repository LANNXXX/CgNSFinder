"""
Neutral Genomic Integration Site Identification Tool

This script performs genome-wide screening to identify suitable neutral integration
sites in Corynebacterium glutamicum based on multiple filtering criteria:

1. Gene Type: Target genes must be annotated as 'hypothetical protein'
2. Neighbor Analysis: At least one neighboring gene (upstream or downstream) must 
   also be a hypothetical protein
3. Length Constraint: Gene length must be within specified range (default: 500-1500 bp)
4. GC Content: Flanking regions (default: 1000 bp on each side) must have GC content 
   within specified range (default: 40-60%)

The tool generates a comprehensive Excel report with color-coded filtering results
for all genes in the genome.

Usage:
    python filterData.py --cds_file <cds_fasta> --fna_file <genome_fasta>
Author: Bioinformatics Research Group
License: The Unlicense (Public Domain)
"""

import argparse
import os
from collections import defaultdict
import re

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from Bio import SeqIO
from Bio.SeqUtils import gc_fraction


# Filter status constants
STATUS_PASSED = "PASSED"
STATUS_PENDING = "Pending"
STATUS_FAILED_NOT_HYPOTHETICAL = "FAILED: Not hypothetical protein"
STATUS_FAILED_NO_HYPOTHETICAL_NEIGHBORS = "FAILED: No hypothetical protein neighbors"
STATUS_FAILED_UPSTREAM_NOT_HYPOTHETICAL = "FAILED: Upstream neighbor not hypothetical"
STATUS_FAILED_DOWNSTREAM_NOT_HYPOTHETICAL = "FAILED: Downstream neighbor not hypothetical"
STATUS_FAILED_BOTH_NEIGHBORS_NOT_HYPOTHETICAL = "FAILED: Both neighbors not hypothetical"
STATUS_FAILED_LENGTH = "FAILED: Length {}bp out of range"
STATUS_FAILED_INSUFFICIENT_FLANKING = "FAILED: Insufficient flanking region"
STATUS_FAILED_GC_CONTENT = "FAILED: GC content out of range"
STATUS_ERROR_GENE_NOT_FOUND = "ERROR: Gene not found in sorted list"


def parse_annotated_fasta(annot_fasta_file):
    """
    Parse an annotated CDS FASTA file from NCBI.
    
    This function extracts gene information including locus tags, products,
    genomic locations, and strand orientation from NCBI-formatted CDS FASTA files.
    
    Args:
        annot_fasta_file (str): Path to the CDS FASTA file containing gene annotations.
        
    Returns:
        dict: A dictionary organized by contig ID, where each contig contains a list
              of gene dictionaries with the following keys:
              - 'id': Locus tag identifier
              - 'locus_tag': Same as id
              - 'contig_id': Contig/chromosome identifier
              - 'products': List of protein product descriptions
              - 'start': 0-based start position
              - 'end': 1-based end position (exclusive)
              - 'strand': 1 for forward strand, -1 for reverse strand
              
    Example:
        >>> genes = parse_annotated_fasta("cds_from_genomic.fna")
        >>> print(genes['NC_003450.3'][0])
        {'id': 'cg0001', 'locus_tag': 'cg0001', ...}
    """
    genes_data = defaultdict(lambda: {
        'products': [], 'locations': [], 'contig_id': None, 'strand': None
    })
    
    for record in SeqIO.parse(annot_fasta_file, "fasta"):
        header = record.description
        
        # Extract contig ID from record ID (format: lcl|CONTIG_cds_...)
        contig_id_match = re.search(r'lcl\|(.+?)_cds', record.id)
        if not contig_id_match:
            continue
        contig_id = contig_id_match.group(1)
        
        # Extract locus tag
        locus_tag_match = re.search(r'\[locus_tag=([^\]]+)\]', header)
        if not locus_tag_match:
            continue
        locus_tag = locus_tag_match.group(1)
        
        # Extract protein product description
        protein_match = re.search(r'\[protein=([^\]]+)\]', header)
        product = protein_match.group(1) if protein_match else "Unknown"
        
        # Extract genomic location and determine strand
        location_match = re.search(r'\[location=([^\]]+)\]', header)
        if not location_match:
            continue
        location_str = location_match.group(1)
        strand = -1 if "complement" in location_str else 1
        
        # Parse start and end coordinates
        coords_match = re.findall(r'(\d+)', location_str)
        if len(coords_match) < 2:
            continue
        start, end = int(coords_match[0]), int(coords_match[1])
        
        # Store gene information
        genes_data[locus_tag]['products'].append(product)
        genes_data[locus_tag]['locations'].append((start, end))
        genes_data[locus_tag]['contig_id'] = contig_id
        if genes_data[locus_tag]['strand'] is None:
            genes_data[locus_tag]['strand'] = strand
    
    # Organize genes by contig for efficient neighbor lookup
    final_genes_by_contig = defaultdict(list)
    for locus_tag, data in genes_data.items():
        min_start = min(loc[0] for loc in data['locations'])
        max_end = max(loc[1] for loc in data['locations'])
        
        gene_obj = {
            'id': locus_tag,
            'locus_tag': locus_tag,
            'contig_id': data['contig_id'],
            'products': data['products'],
            'start': min_start - 1,  # Convert to 0-based
            'end': max_end,
            'strand': data['strand']
        }
        final_genes_by_contig[data['contig_id']].append(gene_obj)
    
    return final_genes_by_contig


def check_gene_type(gene):
    """
    Check if a gene is annotated as a hypothetical protein.
    
    Filtering Criterion 1: A gene qualifies as a neutral site candidate only if
    all its products are annotated as 'hypothetical protein'.
    
    Args:
        gene (dict): Gene dictionary containing 'products' key with list of 
                     protein descriptions.
                     
    Returns:
        bool: True if all products are 'hypothetical protein', False otherwise.
        
    Example:
        >>> gene = {'products': ['hypothetical protein']}
        >>> check_gene_type(gene)
        True
    """
    products = gene.get("products", [])
    if not products:
        return False
    return all(p.lower() == "hypothetical protein" for p in products)


def find_candidate_sites_detailed(genes_by_contig, seq_dict, min_len, max_len, 
                                   flank_len, gc_min, gc_max):
    """
    Screen all genes in the genome and generate detailed filtering reports.
    
    This function applies all filtering criteria sequentially to each gene and
    reports the specific reason for failure or success. GC content is calculated
    for all genes regardless of filtering status.
    
    Args:
        genes_by_contig (dict): Gene dictionary organized by contig ID.
        seq_dict (dict): Dictionary of sequence records indexed by contig ID.
        min_len (int): Minimum acceptable gene length in base pairs.
        max_len (int): Maximum acceptable gene length in base pairs.
        flank_len (int): Length of flanking regions to analyze for GC content.
        gc_min (float): Minimum acceptable GC percentage (0-100).
        gc_max (float): Maximum acceptable GC percentage (0-100).
        
    Returns:
        list: List of dictionaries, each containing detailed information about
              a gene and its filtering status:
              - seq_id: Contig identifier
              - gene_id: Locus tag
              - products: Semicolon-separated protein products
              - start: 1-based start position
              - end: 1-based end position
              - strand: '+' or '-'
              - length: Gene length in bp
              - gc_left(%): GC content of left flanking region
              - gc_right(%): GC content of right flanking region
              - Filter_Status: Filtering outcome (PASSED or FAILED with reason)
              
    Note:
        Filtering is applied in the following order:
        1. Gene type (hypothetical protein)
        2. Neighbor type (at least one hypothetical protein neighbor)
        3. Gene length
        4. Flanking region availability
        5. GC content of flanking regions
    """
    all_gene_reports = []
    
    # Pre-sort all genes by position within each contig for efficient neighbor lookup
    sorted_genes_by_contig = {
        contig: sorted(genes, key=lambda g: g['start'])
        for contig, genes in genes_by_contig.items()
    }
    
    # Flatten gene list for iteration
    all_genes_flat = [gene for genes in genes_by_contig.values() for gene in genes]
    
    for gene in all_genes_flat:
        contig_id = gene['contig_id']
        genome_seq = seq_dict.get(contig_id, {}).seq
        
        # Combine multiple products into a single string
        products_str = "; ".join(gene.get('products', ['N/A']))
        
        # Calculate GC content for flanking regions
        start, end = gene['start'], gene['end']
        left_start = max(0, start - flank_len)
        right_end = min(len(genome_seq), end + flank_len)
        
        gc_left = gc_fraction(genome_seq[left_start:start]) * 100
        gc_right = gc_fraction(genome_seq[end:right_end]) * 100
        
        # Initialize report for this gene
        report = {
            "seq_id": contig_id,
            "gene_id": gene['locus_tag'],
            "products": products_str,
            "start": gene['start'] + 1,  # Convert to 1-based for output
            "end": gene['end'],
            "strand": "+" if gene['strand'] == 1 else "-",
            "length": gene['end'] - gene['start'],
            "gc_left(%)": float(f"{gc_left:.2f}"),
            "gc_right(%)": float(f"{gc_right:.2f}"),
            "Filter_Status": STATUS_PENDING
        }
        
        # --- Apply filtering criteria sequentially ---
        
        # Criterion 1: Gene must be hypothetical protein
        if not check_gene_type(gene):
            report["Filter_Status"] = STATUS_FAILED_NOT_HYPOTHETICAL
            all_gene_reports.append(report)
            continue
        
        # Criterion 2: At least one neighbor must be hypothetical protein
        sorted_gene_list = sorted_genes_by_contig.get(contig_id, [])
        try:
            gene_index = sorted_gene_list.index(gene)
        except ValueError:
            report["Filter_Status"] = STATUS_ERROR_GENE_NOT_FOUND
            all_gene_reports.append(report)
            continue
        
        # Check upstream and downstream neighbors
        is_upstream_hypo = (gene_index > 0 and 
                           check_gene_type(sorted_gene_list[gene_index - 1]))
        is_downstream_hypo = (gene_index < len(sorted_gene_list) - 1 and 
                             check_gene_type(sorted_gene_list[gene_index + 1]))
        
        if not (is_upstream_hypo or is_downstream_hypo):
            if not is_upstream_hypo and not is_downstream_hypo:
                report["Filter_Status"] = STATUS_FAILED_BOTH_NEIGHBORS_NOT_HYPOTHETICAL
            elif not is_upstream_hypo:
                report["Filter_Status"] = STATUS_FAILED_UPSTREAM_NOT_HYPOTHETICAL
            else:
                report["Filter_Status"] = STATUS_FAILED_DOWNSTREAM_NOT_HYPOTHETICAL
            all_gene_reports.append(report)
            continue
        
        # Criterion 3: Gene length must be within specified range
        if not (min_len <= report["length"] <= max_len):
            report["Filter_Status"] = STATUS_FAILED_LENGTH.format(report['length'])
            all_gene_reports.append(report)
            continue
        
        # Criterion 4: Sufficient flanking regions must be available
        if left_start >= start or end >= right_end:
            report["Filter_Status"] = STATUS_FAILED_INSUFFICIENT_FLANKING
            all_gene_reports.append(report)
            continue
        
        # Criterion 5: GC content of flanking regions must be within range
        if not (gc_min <= gc_left <= gc_max and gc_min <= gc_right <= gc_max):
            report["Filter_Status"] = STATUS_FAILED_GC_CONTENT
            all_gene_reports.append(report)
            continue
        
        # All criteria passed
        report["Filter_Status"] = STATUS_PASSED
        all_gene_reports.append(report)
    
    return all_gene_reports


def save_report_to_excel(report_data, output_path):
    """
    Save filtering report to an Excel file with color-coded highlighting.
    
    Genes that pass all filters are highlighted in green with green text.
    Genes that fail any filter are highlighted in yellow with red text.
    
    Args:
        report_data (list): List of gene report dictionaries from 
                           find_candidate_sites_detailed().
        output_path (str): Path where the Excel file should be saved.
        
    Returns:
        None
        
    Note:
        The function will print success/error messages to stdout.
    """
    if not report_data:
        print("No report data generated.")
        return
    
    print(f"\nGenerating Excel report: {output_path}...")
    df = pd.DataFrame(report_data)
    
    # Define color schemes for filtering results
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", 
                             fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", 
                            fill_type="solid")
    red_font = Font(color="9C0006")
    green_font = Font(color="006100")
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Filter_Report"
    
    # Write headers
    ws.append(list(df.columns))
    
    # Write data rows with conditional formatting
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply color scheme based on filter status
        status = row[-1]
        if status == STATUS_PASSED:
            fill = green_fill
            font = green_font
        else:
            fill = yellow_fill
            font = red_font
        
        # Apply fill to all cells in the row
        for c_idx in range(1, len(df.columns) + 1):
            ws.cell(row=r_idx, column=c_idx).fill = fill
        
        # Apply font color to status column
        ws.cell(row=r_idx, column=len(df.columns)).font = font
    
    try:
        wb.save(output_path)
        print(f"Success! Report saved to '{output_path}'.")
    except Exception as e:
        print(f"Error saving Excel file: {e}")


def main(args):
    """
    Main function coordinating file I/O and analysis workflow.
    
    Args:
        args (argparse.Namespace): Command-line arguments containing:
            - fna_file: Path to genome FASTA file
            - cds_file: Path to CDS annotation FASTA file
            - out_file: Path for output Excel report
            - min_len, max_len: Gene length constraints
            - flank_len: Flanking region length for GC analysis
            - gc_min, gc_max: GC content constraints
            
    Returns:
        None
    """
    # Load genome sequence
    try:
        seq_dict = SeqIO.to_dict(SeqIO.parse(args.fna_file, "fasta"))
    except FileNotFoundError:
        print(f"ERROR: Genome file not found -> {args.fna_file}")
        return
    
    # Parse gene annotations
    try:
        genes_by_contig = parse_annotated_fasta(args.cds_file)
    except FileNotFoundError:
        print(f"ERROR: Annotation file not found -> {args.cds_file}")
        return
    
    # Run filtering analysis on all genes
    all_reports = find_candidate_sites_detailed(
        genes_by_contig,
        seq_dict,
        args.min_len,
        args.max_len,
        args.flank_len,
        args.gc_min,
        args.gc_max
    )
    
    # Generate and save detailed Excel report with highlighting
    save_report_to_excel(all_reports, args.out_file)
    
    # Print summary statistics
    final_pass_count = sum(1 for r in all_reports if r["Filter_Status"] == STATUS_PASSED)
    print("\n" + "=" * 50)
    print("FILTERING SUMMARY")
    print("=" * 50)
    print(f"Total genes analyzed: {len(all_reports)}")
    print(f"Sites passing all filters: {final_pass_count}")
    print(f"Pass rate: {final_pass_count/len(all_reports)*100:.2f}%" if all_reports else "N/A")
    print("=" * 50)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Identify neutral genomic integration sites from genome and "
                    "CDS annotation FASTA files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Example usage:
    # Basic usage with default parameters
    python filterData.py --cds_file cds.fna --fna_file genome.fna
    
    # Custom filtering parameters
    python filterData.py --cds_file cds.fna --fna_file genome.fna \\
        --min_len 600 --max_len 1200 --gc_min 45 --gc_max 55
        
Filtering Criteria:
    1. Gene must be annotated as 'hypothetical protein'
    2. At least one neighbor must also be hypothetical protein
    3. Gene length must be within min_len to max_len range
    4. Flanking regions (flank_len bp) must have GC% within gc_min to gc_max
        """
    )
    
    # Input files
    parser.add_argument(
        "--cds_file",
        default="./datas/cds_from_genomic.fna",
        help="Path to CDS FASTA file with gene annotations (default: ./datas/cds_from_genomic.fna)"
    )
    parser.add_argument(
        "--fna_file",
        default="./datas/GCF_000011325.1_ASM1132v1_genomic.fna",
        help="Path to genome sequence FASTA file (default: ./datas/GCF_000011325.1_ASM1132v1_genomic.fna)"
    )
    
    # Output file
    parser.add_argument(
        "--out_file",
        default="genome_screening_report.xlsx",
        help="Path for output Excel report (default: genome_screening_report.xlsx)"
    )
    
    # Filtering parameters
    parser.add_argument(
        "--min_len",
        type=int,
        default=500,
        help="Minimum gene length in base pairs (default: 500)"
    )
    parser.add_argument(
        "--max_len",
        type=int,
        default=1500,
        help="Maximum gene length in base pairs (default: 1500)"
    )
    parser.add_argument(
        "--flank_len",
        type=int,
        default=1000,
        help="Length of flanking regions for GC analysis in bp (default: 1000)"
    )
    parser.add_argument(
        "--gc_min",
        type=float,
        default=40.0,
        help="Minimum GC content percentage for flanking regions (default: 40.0)"
    )
    parser.add_argument(
        "--gc_max",
        type=float,
        default=60.0,
        help="Maximum GC content percentage for flanking regions (default: 60.0)"
    )
    
    args = parser.parse_args()
    main(args)
